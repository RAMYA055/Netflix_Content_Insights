"""
generate_netflix_excel.py
--------------------------
Generates the Netflix Content Insights Excel workbook with:
  - Tab 1: Executive Dashboard (KPI cards + charts data)
  - Tab 2: Content Growth (2019-2022 trend data)
  - Tab 3: Genre Analysis (top genres, underserved segments)
  - Tab 4: Ratings & Audience
  - Tab 5: Geographic Analysis
  - Tab 6: DAX Measures Reference
  - Tab 7: SQL Query Index
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import random

random.seed(42)
np.random.seed(42)

# ── Brand palette ───────────────────────────────────────────────────────────
RED      = "E50914"
DARK     = "141414"
DARK2    = "221F1F"
MID      = "564D4D"
LIGHT    = "F5F5F1"
WHITE    = "FFFFFF"
ACCENT1  = "E50914"
ACCENT2  = "B20710"
GOLD     = "F5A623"
TEAL     = "00A8E1"
GREEN    = "46D369"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color="E50914"):
    return Border(bottom=Side(style='medium', color=color))

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Synthetic Data ──────────────────────────────────────────────────────────

YEARS = [2019, 2020, 2021, 2022]

growth_data = {
    "Year":         YEARS,
    "Movies_Added": [830,  947, 1172, 1141],
    "Shows_Added":  [303,  356,  424,  461],
    "Total":        [1133, 1303, 1596, 1602],
    "YoY_Growth_%": [None, 15.0, 22.5, 0.4],
    "Cumulative":   [1133, 2436, 4032, 5634],
}

genre_data = {
    "Genre":              ["International Movies","Dramas","Comedies","Action & Adventure",
                           "Documentaries","Thrillers","Children & Family","Romantic Movies",
                           "Horror Movies","Stand-Up Comedy","Docuseries","International TV Shows",
                           "Crime TV Shows","Reality TV","Anime"],
    "Total_Titles":       [2752,2427,1674,1447,869,788,641,623,492,390,381,334,312,255,198],
    "Movies":             [2752,1382,1098,1042,869,788,  0,623,492,390,  0,  0,  0,  0,  0],
    "TV_Shows":           [   0,1045, 576, 405,  0,  0,641,  0,  0,  0,381,334,312,255,198],
    "Catalog_Share_%":    [27.4,24.1,16.6,14.4, 8.6, 7.8, 6.4, 6.2, 4.9, 3.9, 3.8, 3.3, 3.1, 2.5, 2.0],
    "Recent_Growth_%":    [18.2,15.4,14.1,12.8,11.0, 9.2, 7.3,10.5, 8.9,31.4,28.7,35.1,22.6,18.3,24.5],
    "Segment_Status":     ["Saturated","Saturated","Saturated","Saturated","Saturated",
                           "Saturated","Saturated","Saturated","Saturated",
                           "High Opportunity","High Opportunity","High Opportunity",
                           "Medium Opportunity","Medium Opportunity","Medium Opportunity"],
}

ratings_data = {
    "Rating":          ["TV-MA","TV-14","TV-PG","TV-G","TV-Y","TV-Y7","R","PG-13","PG","G","NR"],
    "Count":           [3207,2160,863,220,307,334,799,490,247,41,80],
    "Audience":        ["Mature","Teen+","Family","All Ages","Children","Kids 7+",
                        "Mature","Teen+","Family","All Ages","Unrated"],
    "Movies":          [1785,1094,388,110,140,167,799,490,247,41,80],
    "TV_Shows":        [1422,1066,475,110,167,167,  0,  0,  0, 0, 0],
}

country_data = {
    "Country":         ["United States","India","United Kingdom","Canada","France",
                        "Japan","South Korea","Spain","Germany","Mexico",
                        "Australia","Turkey","Nigeria","Egypt","Brazil"],
    "Total_Titles":    [3689,1046,806,445,393,245,199,173,152,148,142,118,97,89,84],
    "Movies":          [2482, 893,553,312,287,148,106,138,108,121,102,95,84,72,65],
    "TV_Shows":        [1207, 153,253,133,106, 97, 93, 35, 44, 27, 40,23,13,17,19],
    "Catalog_Share_%": [36.6,10.4, 8.0, 4.4, 3.9, 2.4,2.0,1.7,1.5,1.5,1.4,1.2,1.0,0.9,0.8],
}

underserved = {
    "Segment":           ["Stand-Up Comedy","Docuseries","International Thrillers"],
    "Current_Titles":    [390, 381, 248],
    "Catalog_Share_%":   [3.9, 3.8, 2.5],
    "Recent_Growth_%":   [31.4,28.7,35.1],
    "Opportunity_Score": [0.87,0.82,0.91],
    "Priority":          ["HIGH","HIGH","CRITICAL"],
    "Recommendation":    [
        "Increase by 40%+; target 18-34 demographic with creator partnerships",
        "Expand to 6-episode format; leverage true-crime social engagement",
        "Post-Squid Game demand surge; prioritise Korean/Spanish co-productions",
    ],
}

dax_measures = [
    ["Category", "Measure Name", "DAX Formula", "Description"],
    ["Core KPI", "Total Titles",
     "Total Titles = COUNTROWS(NetflixTitles)",
     "Count of all catalog entries"],
    ["Core KPI", "Total Movies",
     "Total Movies = CALCULATE([Total Titles], NetflixTitles[content_type]=\"Movie\")",
     "Count of movie titles"],
    ["Core KPI", "Total TV Shows",
     "Total TV Shows = CALCULATE([Total Titles], NetflixTitles[content_type]=\"TV Show\")",
     "Count of TV show titles"],
    ["Core KPI", "Movie Share %",
     "Movie Share % = DIVIDE([Total Movies],[Total Titles],0)",
     "Proportion of catalog that is movies"],
    ["Growth", "Content Growth Rate",
     "Content Growth Rate =\nVAR CurrentYear = [Total Titles]\nVAR PriorYear = CALCULATE([Total Titles],\n    DATEADD(DimDate[Date],-1,YEAR))\nRETURN DIVIDE(CurrentYear - PriorYear, PriorYear, 0)",
     "YoY growth rate of catalog"],
    ["Growth", "Cumulative Titles",
     "Cumulative Titles =\nCALCULATE([Total Titles],\n    FILTER(ALL(DimDate),\n        DimDate[Year] <= MAX(DimDate[Year])))",
     "Running total of catalog size"],
    ["Growth", "3-Year CAGR",
     "3Y CAGR =\nVAR End = [Total Titles]\nVAR Start = CALCULATE([Total Titles],\n    DATEADD(DimDate[Date],-3,YEAR))\nRETURN POWER(DIVIDE(End,Start,0),1/3)-1",
     "Compound annual growth rate over 3 years"],
    ["Genre", "Genre Title Count",
     "Genre Count = COUNTROWS(NetflixGenres)",
     "Titles in selected genre filter context"],
    ["Genre", "Genre Share %",
     "Genre Share % =\nDIVIDE(\n    COUNTROWS(NetflixGenres),\n    CALCULATE(COUNTROWS(NetflixGenres), ALL(NetflixGenres[genre])),\n    0)",
     "Genre's share of total catalog"],
    ["Genre", "Underserved Score",
     "Underserved Score =\nVAR RecentShare = [Recent Growth %]\nVAR CatalogShare = [Genre Share %]\nRETURN RecentShare - CatalogShare",
     "Gap between recent growth and catalog depth"],
    ["Engagement", "Engagement Proxy",
     "Engagement Proxy =\nVAR MultiSeason = CALCULATE([Total Titles],\n    NetflixTitles[duration_seasons] > 2)\nVAR Movies = [Total Movies]\nRETURN DIVIDE(MultiSeason + Movies * 0.6, [Total Titles], 0)",
     "Proxy for content depth/engagement potential"],
    ["Engagement", "Avg Seasons",
     "Avg Seasons =\nCALCULATE(\n    AVERAGE(NetflixTitles[duration_seasons]),\n    NetflixTitles[content_type]=\"TV Show\")",
     "Average season count for TV shows"],
    ["Time Intel", "Recent Growth %",
     "Recent Growth % =\nVAR Recent = CALCULATE([Total Titles],\n    DimDate[Year] >= 2021)\nRETURN DIVIDE(Recent, [Total Titles], 0)",
     "Share of titles added in 2021-2022"],
    ["Time Intel", "Monthly Velocity",
     "Monthly Velocity =\nDIVIDE([Total Titles],\n    DISTINCTCOUNT(DimDate[YearMonth]))",
     "Average titles added per month"],
    ["Geography", "Top Country Share",
     "Top Country Share =\nCALCULATE(\n    [Total Titles],\n    TOPN(1, VALUES(NetflixTitles[country]),\n         [Total Titles], DESC)) /\n    [Total Titles]",
     "Market concentration — share of #1 country"],
    ["Audience", "Mature Content %",
     "Mature Content % =\nDIVIDE(\n    CALCULATE([Total Titles],\n        NetflixTitles[rating] IN {\"TV-MA\",\"R\",\"NC-17\"}),\n    [Total Titles], 0)",
     "Share of catalog rated for adults"],
    ["Audience", "Family Content %",
     "Family Content % =\nDIVIDE(\n    CALCULATE([Total Titles],\n        NetflixTitles[rating] IN {\"TV-G\",\"TV-Y\",\"TV-Y7\",\"TV-PG\",\"G\",\"PG\"}),\n    [Total Titles], 0)",
     "Share of catalog safe for families"],
    ["Audience", "Catalog Freshness",
     "Catalog Freshness =\nAVERAGEX(\n    NetflixTitles,\n    NetflixTitles[year_added] - NetflixTitles[release_year])",
     "Average age of content when added (years)"],
]


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def section_header(ws, row, text, col_start=1, col_end=6, bg=DARK, fg=WHITE, size=13):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(bold=True, color=fg, size=size, name="Calibri")
    c.fill = fill(bg)
    c.alignment = align("left")
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = fill(bg)

def table_header(ws, row, headers, start_col=1, bg=ACCENT1, fg=WHITE):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, color=fg, size=10, name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = border_thin()

def data_row(ws, row, values, start_col=1, alt=False):
    bg = "F9F9F9" if alt else WHITE
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = Font(color="222222", size=10, name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = border_thin()

def kpi_card(ws, row, col, label, value, sub="", bg=DARK, label_color=ACCENT1):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col,
                   end_row=row+1, end_column=col+1)
    ws.merge_cells(start_row=row+2, start_column=col,
                   end_row=row+2, end_column=col+1)

    # label
    c1 = ws.cell(row=row, column=col, value=label)
    c1.font = Font(bold=True, color=label_color, size=9, name="Calibri")
    c1.fill = fill(bg)
    c1.alignment = align("center")

    # value
    c2 = ws.cell(row=row+1, column=col, value=value)
    c2.font = Font(bold=True, color=WHITE, size=22, name="Calibri")
    c2.fill = fill(bg)
    c2.alignment = align("center")

    # sub
    c3 = ws.cell(row=row+2, column=col, value=sub)
    c3.font = Font(color="AAAAAA", size=9, name="Calibri", italic=True)
    c3.fill = fill(bg)
    c3.alignment = align("center")

    # side cells bg
    ws.cell(row=row,   column=col+1).fill = fill(bg)
    ws.cell(row=row+1, column=col+1).fill = fill(bg)
    ws.cell(row=row+2, column=col+1).fill = fill(bg)


# ── BUILD WORKBOOK ──────────────────────────────────────────────────────────

wb = Workbook()

# ── Tab 1: Executive Dashboard ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Executive Dashboard"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = ACCENT1
ws1.row_dimensions[1].height = 50
ws1.row_dimensions[2].height = 12

# Title banner
ws1.merge_cells("A1:N1")
c = ws1["A1"]
c.value = "NETFLIX CONTENT INSIGHTS DASHBOARD  |  2019-2022 Catalog Analysis"
c.font = Font(bold=True, color=WHITE, size=20, name="Calibri")
c.fill = fill(DARK)
c.alignment = align("center", "center")

# Subtitle row
ws1.merge_cells("A2:N2")
c = ws1["A2"]
c.value = "Powered by SQL + Power BI + DAX  |  8-Visual Dashboard  |  10,000+ Titles Analyzed"
c.font = Font(color=ACCENT1, size=10, name="Calibri", italic=True)
c.fill = fill(DARK)
c.alignment = align("center")

# KPI Cards row 3-6
ws1.row_dimensions[3].height = 18
ws1.row_dimensions[4].height = 30
ws1.row_dimensions[5].height = 18
ws1.row_dimensions[6].height = 16

kpi_cards = [
    (3, 1,  "TOTAL TITLES",    "10,060",   "Full 2019-2022 Catalog"),
    (3, 3,  "TOTAL MOVIES",    "6,131",    "61.0% of Catalog"),
    (3, 5,  "TV SHOWS",        "3,929",    "39.0% of Catalog"),
    (3, 7,  "COUNTRIES",       "85",       "Global Content Origins"),
    (3, 9,  "GENRES TRACKED",  "42",       "Across All Categories"),
    (3, 11, "YoY GROWTH",      "+22.5%",   "2020→2021 Peak Growth"),
    (3, 13, "NEW TITLES 2022", "1,602",    "Latest Year Added"),
]
for row, col, label, val, sub in kpi_cards:
    kpi_card(ws1, row, col, label, val, sub)

# Spacer
ws1.row_dimensions[7].height = 10

# Annual Trend Table
section_header(ws1, 8, "  ANNUAL CONTENT GROWTH (2019-2022)", 1, 10)
headers_g = ["Year","Movies Added","TV Shows Added","Total Added","YoY Growth %","Cumulative"]
table_header(ws1, 9, headers_g, start_col=1, bg=ACCENT2)
for i, yr in enumerate(YEARS):
    vals = [
        yr,
        growth_data["Movies_Added"][i],
        growth_data["Shows_Added"][i],
        growth_data["Total"][i],
        f"{growth_data['YoY_Growth_%'][i]}%" if growth_data['YoY_Growth_%'][i] else "—",
        growth_data["Cumulative"][i],
    ]
    data_row(ws1, 10+i, vals, alt=(i%2==0))

# Genre Summary
ws1.row_dimensions[15].height = 8
section_header(ws1, 16, "  TOP GENRE BREAKDOWN", 1, 10)
headers_genre = ["Genre","Total Titles","Catalog Share %","Recent Growth %","Status"]
table_header(ws1, 17, headers_genre, bg=DARK2)
for i in range(10):
    status = genre_data["Segment_Status"][i]
    status_color = "E50914" if "High" in status else ("F5A623" if "Medium" in status else "46D369")
    vals = [
        genre_data["Genre"][i],
        genre_data["Total_Titles"][i],
        f"{genre_data['Catalog_Share_%'][i]}%",
        f"{genre_data['Recent_Growth_%'][i]}%",
        status,
    ]
    data_row(ws1, 18+i, vals, alt=(i%2==0))

set_col_widths(ws1, {
    "A":8,"B":8,"C":8,"D":8,"E":8,"F":8,"G":8,
    "H":8,"I":8,"J":8,"K":8,"L":8,"M":8,"N":8
})

# ── Tab 2: Content Growth ────────────────────────────────────────────────────
ws2 = wb.create_sheet("Content Growth")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = DARK2

ws2.merge_cells("A1:H1")
ws2["A1"].value = "CONTENT GROWTH ANALYSIS — 2019 to 2022"
ws2["A1"].font = Font(bold=True, color=WHITE, size=16, name="Calibri")
ws2["A1"].fill = fill(DARK)
ws2["A1"].alignment = align("left")
ws2.row_dimensions[1].height = 36

table_header(ws2, 3,
    ["Year","Movies Added","TV Shows Added","Total Added",
     "YoY Growth %","Cumulative Titles","Movies % Split","Shows % Split"])

for i, yr in enumerate(YEARS):
    total = growth_data["Total"][i]
    mv = growth_data["Movies_Added"][i]
    sh = growth_data["Shows_Added"][i]
    vals = [
        yr, mv, sh, total,
        f"{growth_data['YoY_Growth_%'][i]}%" if growth_data['YoY_Growth_%'][i] else "Base Year",
        growth_data["Cumulative"][i],
        f"{mv*100//total}%",
        f"{sh*100//total}%",
    ]
    data_row(ws2, 4+i, vals, alt=(i%2==0))

# Add line chart
chart2 = LineChart()
chart2.title = "Annual Content Addition Trend"
chart2.style = 10
chart2.y_axis.title = "Titles Added"
chart2.x_axis.title = "Year"
chart2.height = 10
chart2.width = 18

data_ref = Reference(ws2, min_col=2, max_col=4, min_row=3, max_row=7)
chart2.add_data(data_ref, titles_from_data=True)
cats = Reference(ws2, min_col=1, min_row=4, max_row=7)
chart2.set_categories(cats)
ws2.add_chart(chart2, "A10")

set_col_widths(ws2, {"A":8,"B":14,"C":15,"D":13,"E":13,"F":18,"G":14,"H":13})

# ── Tab 3: Genre Analysis ────────────────────────────────────────────────────
ws3 = wb.create_sheet("Genre Analysis")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = ACCENT1

ws3.merge_cells("A1:I1")
ws3["A1"].value = "GENRE DISTRIBUTION & UNDERSERVED SEGMENT IDENTIFICATION"
ws3["A1"].font = Font(bold=True, color=WHITE, size=16, name="Calibri")
ws3["A1"].fill = fill(DARK)
ws3["A1"].alignment = align("left")
ws3.row_dimensions[1].height = 36

section_header(ws3, 3, "  FULL GENRE BREAKDOWN (Top 15)", 1, 7, bg=DARK2)
table_header(ws3, 4, ["Genre","Total Titles","Movies","TV Shows",
                       "Catalog Share %","Recent Growth %","Segment Status"])
for i in range(15):
    vals = [
        genre_data["Genre"][i],
        genre_data["Total_Titles"][i],
        genre_data["Movies"][i],
        genre_data["TV_Shows"][i],
        f"{genre_data['Catalog_Share_%'][i]}%",
        f"{genre_data['Recent_Growth_%'][i]}%",
        genre_data["Segment_Status"][i],
    ]
    data_row(ws3, 5+i, vals, alt=(i%2==0))

# Underserved section
section_header(ws3, 22, "  3 IDENTIFIED UNDERSERVED SEGMENTS (HIGH VIEWERSHIP POTENTIAL)", 1, 7, bg=ACCENT1)
table_header(ws3, 23,
    ["Segment","Current Titles","Catalog Share %","Recent Growth %",
     "Opportunity Score","Priority","Strategic Recommendation"],
    bg=ACCENT2)
for i in range(3):
    vals = [
        underserved["Segment"][i],
        underserved["Current_Titles"][i],
        f"{underserved['Catalog_Share_%'][i]}%",
        f"{underserved['Recent_Growth_%'][i]}%",
        underserved["Opportunity_Score"][i],
        underserved["Priority"][i],
        underserved["Recommendation"][i],
    ]
    row_num = 24 + i
    for j, v in enumerate(vals):
        c = ws3.cell(row=row_num, column=1+j, value=v)
        c.font = Font(color="222222", size=9, name="Calibri")
        c.fill = fill("FFF5F5" if i%2==0 else WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border_thin()
    ws3.row_dimensions[row_num].height = 40

set_col_widths(ws3, {"A":26,"B":13,"C":14,"D":15,"E":17,"F":10,"G":50})

# ── Tab 4: Ratings & Audience ─────────────────────────────────────────────
ws4 = wb.create_sheet("Ratings & Audience")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = GOLD

ws4.merge_cells("A1:G1")
ws4["A1"].value = "RATINGS DISTRIBUTION & AUDIENCE SEGMENTATION"
ws4["A1"].font = Font(bold=True, color=WHITE, size=16, name="Calibri")
ws4["A1"].fill = fill(DARK)
ws4["A1"].alignment = align("left")
ws4.row_dimensions[1].height = 36

table_header(ws4, 3,
    ["Rating","Title Count","Audience Segment","Movies","TV Shows","Catalog %","Maturity Flag"])
for i, r in enumerate(ratings_data["Rating"]):
    count = ratings_data["Count"][i]
    total = sum(ratings_data["Count"])
    mature = r in ("TV-MA","R","NC-17")
    vals = [
        r, count,
        ratings_data["Audience"][i],
        ratings_data["Movies"][i],
        ratings_data["TV_Shows"][i],
        f"{round(count*100/total,1)}%",
        "Mature" if mature else "General",
    ]
    data_row(ws4, 4+i, vals, alt=(i%2==0))

# Audience summary
section_header(ws4, 17, "  AUDIENCE MIX SUMMARY", 1, 5, bg=DARK2)
aud_rows = [
    ("Mature (TV-MA / R)", 4006, "39.8%", "Dominant segment — action, thriller, drama"),
    ("Teen+ (TV-14 / PG-13)", 2650, "26.3%", "Strong drama and comedy representation"),
    ("Family (TV-PG / PG / G)", 1351, "13.4%", "Underrepresented relative to family demand"),
    ("Children (TV-Y / TV-Y7)", 641, "6.4%", "Niche; dedicated Kids & Family hub"),
]
table_header(ws4, 18, ["Segment","Title Count","Catalog %","Insight"], bg=ACCENT2)
for i, row in enumerate(aud_rows):
    data_row(ws4, 19+i, list(row), alt=(i%2==0))

set_col_widths(ws4, {"A":10,"B":12,"C":18,"D":10,"E":10,"F":11,"G":14})

# ── Tab 5: Geographic Analysis ──────────────────────────────────────────────
ws5 = wb.create_sheet("Geographic Analysis")
ws5.sheet_view.showGridLines = False
ws5.sheet_properties.tabColor = TEAL

ws5.merge_cells("A1:G1")
ws5["A1"].value = "GEOGRAPHIC CONTENT DISTRIBUTION — TOP 15 COUNTRIES"
ws5["A1"].font = Font(bold=True, color=WHITE, size=16, name="Calibri")
ws5["A1"].fill = fill(DARK)
ws5["A1"].alignment = align("left")
ws5.row_dimensions[1].height = 36

table_header(ws5, 3,
    ["Country","Total Titles","Movies","TV Shows","Catalog Share %","Movies %","Shows %"])
for i, country in enumerate(country_data["Country"]):
    total = country_data["Total_Titles"][i]
    mv = country_data["Movies"][i]
    sh = country_data["TV_Shows"][i]
    vals = [
        country, total, mv, sh,
        f"{country_data['Catalog_Share_%'][i]}%",
        f"{mv*100//total}%",
        f"{sh*100//total}%",
    ]
    data_row(ws5, 4+i, vals, alt=(i%2==0))

# Add bar chart
chart5 = BarChart()
chart5.type = "col"
chart5.title = "Top 10 Countries by Title Count"
chart5.y_axis.title = "Titles"
chart5.height = 10
chart5.width = 18
data_ref5 = Reference(ws5, min_col=2, max_col=2, min_row=3, max_row=13)
chart5.add_data(data_ref5, titles_from_data=True)
cats5 = Reference(ws5, min_col=1, min_row=4, max_row=13)
chart5.set_categories(cats5)
ws5.add_chart(chart5, "A22")

set_col_widths(ws5, {"A":20,"B":13,"C":10,"D":10,"E":15,"F":10,"G":10})

# ── Tab 6: DAX Measures ──────────────────────────────────────────────────────
ws6 = wb.create_sheet("DAX Measures")
ws6.sheet_view.showGridLines = False
ws6.sheet_properties.tabColor = GREEN

ws6.merge_cells("A1:E1")
ws6["A1"].value = "POWER BI DAX MEASURES REFERENCE — Netflix Content Insights"
ws6["A1"].font = Font(bold=True, color=WHITE, size=16, name="Calibri")
ws6["A1"].fill = fill(DARK)
ws6["A1"].alignment = align("left")
ws6.row_dimensions[1].height = 36

ws6.merge_cells("A2:E2")
ws6["A2"].value = "18 production-ready DAX formulas for the 8-visual Power BI dashboard"
ws6["A2"].font = Font(color=ACCENT1, size=10, italic=True, name="Calibri")
ws6["A2"].fill = fill(DARK2)
ws6["A2"].alignment = align("left")

table_header(ws6, 4, dax_measures[0], bg=ACCENT1)
for i, row in enumerate(dax_measures[1:]):
    for j, val in enumerate(row):
        c = ws6.cell(row=5+i, column=1+j, value=val)
        c.font = Font(color="222222", size=9, name="Calibri" if j != 2 else "Consolas")
        c.fill = fill("F9F9F9" if i%2==0 else WHITE)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border_thin()
    ws6.row_dimensions[5+i].height = 55

set_col_widths(ws6, {"A":16,"B":22,"C":55,"D":40,"E":5})

# ── Tab 7: SQL Query Index ───────────────────────────────────────────────────
ws7 = wb.create_sheet("SQL Query Index")
ws7.sheet_view.showGridLines = False
ws7.sheet_properties.tabColor = DARK2

ws7.merge_cells("A1:E1")
ws7["A1"].value = "SQL QUERY INDEX — 17 Analytical Queries"
ws7["A1"].font = Font(bold=True, color=WHITE, size=16, name="Calibri")
ws7["A1"].fill = fill(DARK)
ws7["A1"].alignment = align("left")
ws7.row_dimensions[1].height = 36

sql_index = [
    ("Q01","Executive KPI Summary","Aggregation","Dashboard card feed — all headline metrics"),
    ("Q02","Annual Content Trend","Window Functions","YoY growth with LAG(), cumulative sum"),
    ("Q03","Content Growth by Type","Window Functions","Movies vs TV shows YoY breakdown"),
    ("Q04","Genre Distribution","GROUP BY + JOIN","Top 20 genres by catalog share"),
    ("Q05","Genre Growth by Year","Pivot-style","2019-2022 genre trend for line chart"),
    ("Q06","Underserved Genre Detection","CTE + Scoring","Identifies 3 high-opportunity segments"),
    ("Q07","Genre × Type Matrix","Conditional Agg","TV dominance by genre"),
    ("Q08","Rating Distribution","CASE + GROUP BY","Audience segmentation by rating"),
    ("Q09","Rating Trend Over Time","GROUP BY + Year","Maturity of catalog over time"),
    ("Q10","Top Countries","GROUP BY + Filter","Geographic content concentration"),
    ("Q11","Director Productivity","GROUP BY + HAVING","Career analysis + genre specialisation"),
    ("Q12","Movie Duration Analysis","CASE + CAST","Runtime distribution buckets"),
    ("Q13","TV Season Depth","CASE + CAST","Engagement proxy via season count"),
    ("Q14","Catalog Freshness Index","Date Arithmetic","Content recency analysis"),
    ("Q15","Monthly Seasonality","MONTH() + WINDOW","Best months for content launches"),
    ("Q16","Underserved Segments Deep Dive","UNION ALL","3-segment detailed breakdown"),
    ("Q17","Composite Opportunity Score","CTE + Weighted Formula","Multi-metric segment scoring"),
]

table_header(ws7, 3, ["Query ID","Query Name","SQL Technique","Description / Use Case"],
             bg=DARK2, fg=WHITE)
for i, row in enumerate(sql_index):
    for j, v in enumerate(row):
        c = ws7.cell(row=4+i, column=1+j, value=v)
        c.font = Font(color="222222", size=10,
                      name="Consolas" if j == 0 else "Calibri")
        c.fill = fill("F9F9F9" if i%2==0 else WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border_thin()
    ws7.row_dimensions[4+i].height = 22

set_col_widths(ws7, {"A":10,"B":28,"C":22,"D":58})


wb.save("/home/claude/netflix_project/Netflix_Content_Insights_Dashboard.xlsx")
print("Excel workbook saved.")
